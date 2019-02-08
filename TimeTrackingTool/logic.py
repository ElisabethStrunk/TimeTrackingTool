import time
import enum
from openpyxl import *
from openpyxl.styles import Font

class Timer:
  def __init__(self):
    self.start = time.time()

  def restart(self):
    self.start = time.time()

  def get_time_hhmmss(self):
    end = time.time()
    m, s = divmod(end - self.start, 60)
    h, m = divmod(m, 60)
    time_str = "%02d:%02d:%02d" % (h, m, s)
    return time_str


class State(enum.Enum):
    beginning = 1
    started = 2
    stopped = 3


class Logging:
    def __init__(self, path):
        self.path = path
        try:
            # load excel file
            self.wb = load_workbook(self.path)
            self.sheet = self.wb.worksheets[0]
            # determine first empty row
            for cell in self.sheet["A"]:
                if cell.value is None:
                    self.starting_row = cell.row
                    break
            else:
                self.starting_row = cell.row + 1
        except:
            # create excel file
            self.wb = Workbook()
            self.sheet = self.wb.worksheets[0]
            self.sheet.title = "Time Tracker"
            # write header in ecxel
            self.write(1, 1, "Date")
            self.write(1, 2, "Time")
            self.write(1, 3, "Task")
            # Make header bold
            for cell in self.sheet["1:1"]:
                cell.font = Font(bold=True)
            # set first empty row to row number 2
            self.starting_row = 2
            # save changes
            self.save()

    def write(self, row, column, content):
        self.sheet.cell(row=row, column=column).value = content

    def read(self, row, column):
        return self.sheet.cell(row=row, column=column).value

    def save(self):
        self.wb.save(self.path)

    def write_new_entry(self, date, time, task):
        self.write(self.starting_row, 1, date)
        self.write(self.starting_row, 2, time)
        self.write(self.starting_row, 3, task)
        self.starting_row += 1